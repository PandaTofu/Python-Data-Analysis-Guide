from pandas import DataFrame

df = DataFrame([(1,2,3), (4,5,6)], columns=['x','y','z'])

df

from openpyxl import load_workbook

wb = load_workbook('examples/example_6_10.xlsx')

ws = wb.create_sheet('my_sheet')

ws.append(list(df.columns))

for row in df.values:
    ws.append(list(row))

wb.save('examples/example_6_10.xlsx')

from openpyxl import Workbook

wb = Workbook()

