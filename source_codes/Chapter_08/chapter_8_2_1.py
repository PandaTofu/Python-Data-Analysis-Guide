from openpyxl import load_workbook

wb = load_workbook('examples/example_6_10.xlsx')

ws = wb['Sheet1']

from openpyxl import load_workbook

wb = load_workbook('examples/example_6_10.xlsx')

ws = wb['Sheet1']

data = list(ws.values)

data

from pandas import DataFrame

header, values = data[0], data[1:]

df = DataFrame(values, columns=header)

df

